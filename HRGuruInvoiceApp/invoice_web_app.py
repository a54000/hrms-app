from __future__ import annotations

import base64
import hmac
import json
import math
import os
import re
import struct
import zlib
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from textwrap import wrap

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from flask import Blueprint, Flask, Response, redirect, render_template_string, request, send_file, url_for
from werkzeug.middleware.proxy_fix import ProxyFix
import fitz


APP_ROOT = Path(__file__).resolve().parent
LEGACY_OUTPUT_DIR = APP_ROOT / "generated_invoices"
OUTPUT_DIR = Path(os.environ.get("INVOICE_OUTPUT_DIR", r"D:\HRGURU\Invoices\Invoices"))
MIS_STORE = APP_ROOT / "invoice_mis.json"
REFERENCES_STORE = APP_ROOT / "invoice_references.json"
CLIENTS_STORE = APP_ROOT / "invoice_clients.json"
MIS_EXCEL_PATH = APP_ROOT / "Invoice_MIS_Export.xlsx"
LOGO_PATH = APP_ROOT / "invoice_static" / "logo.png"
PDF_LOGO_PATH = APP_ROOT / "invoice_static" / "logo.jpg"

COMPANY = {
    "name": "HR Guru Placement Services Pvt Ltd",
    "address": "1202, Tower -8, Orchid Petals, Sector - 49, Gurgaon, Haryana, 122018",
    "contact": "+91 99719 33995",
    "gstin": "06AAJCC5251B1Z2",
    "pan": "AAJCC5251B",
    "state": "Haryana",
    "state_code": "06",
    "bank": "IndusInd Bank",
    "account": "250406202101",
    "ifsc": "INDB0000316",
    "branch": "Omaxe City Centre, Sector-49, Gurgaon",
}

CLIENTS = {
    "EY-Hyd": {
        "name": "Ernst & Young LLP",
        "address": "18th Floor, The Skyview 10, South Lobby, Survey No 83/1, Raidurgam, Hyderabad, Telangana-500032",
        "gstin": "36AAEFE1763C1ZT",
        "pan": "AAEFE1763C",
        "state": "Telangana",
    },
    "EY-GGN": {
        "name": "Ernst & Young LLP",
        "address": "Ground Floor, Plot No 67, Institutional Area, Sector-44, Gurugram, Haryana-122003",
        "gstin": "06AAEFE1763C1ZW",
        "pan": "AAEFE1763C",
        "state": "Haryana",
    },
    "EY-Pune": {
        "name": "Ernst & Young LLP",
        "address": "Ground Floor, Tower C, Tech Park One, Yerwada, Pune, Maharashtra-411006",
        "gstin": "27AAEFE1763C1ZS",
        "pan": "AAEFE1763C",
        "state": "Maharashtra",
    },
    "EY-Blore": {
        "name": "Ernst & Young LLP",
        "address": "UB City Canberra Block, No. 24, Vittal Mallya Road, Bengaluru, Karnataka-560001",
        "gstin": "29AAEFE1763C2ZN",
        "pan": "AAEFE1763C",
        "state": "Karnataka",
    },
    "EY-Mumbai": {
        "name": "Ernst & Young LLP",
        "address": "The Ruby 29, Senapati Bapat Marg, Dadar West, Mumbai-400028, Maharashtra",
        "gstin": "27AAEFE1763C1ZS",
        "pan": "AAEFE1763C",
        "state": "Maharashtra",
    },
    "Taggd": {
        "name": "Talent Hired-The Job Store Private Limited",
        "address": "Plot No. A-10, Infocity, Phase-1, Sector 34, Gurugram, Haryana-122001",
        "gstin": "06AAECT4240J1ZE",
        "pan": "AAECT4240J1",
        "state": "Haryana",
        "state_code": "07",
        "buyer_po": "THPO/00525",
        "hsn_sac": "998512",
        "spoc": "Amit Garg",
    },
    "H&B": {
        "name": "HAVER & BOECKER INDIA Pvt. Ltd",
        "address": "Survey No. 32/4/41 & 42 Khandiwada, Baroda Halol Road, Post Asoj, Vadodara 391510 Gujarat",
        "gstin": "24AABCH9243A1Z1",
        "pan": "AABCH9243A",
        "state": "Gujarat",
    },
    "Triam": {
        "name": "TRIAM SECURITY (INDIA) PRIVATE LIMITED",
        "address": "208, Golden Park Society, Ashram Road, Nr. Nav Gujarat College of Computer Application, Usmanpura, Ahmedabad-380013",
        "gstin": "24AALCT1625Q1ZW",
        "pan": "AALCT1625Q",
        "state": "Gujarat",
    },
}

REFERENCES = ["Pooja Lachhwani", "Amit Garg", "Surinder Singh", "HR Guru", "Client Referral"]

BASE_MIS = [
    {"name": "Supriya Balasaheb Sapkal", "client": "EY-GGN", "role": "Senior Consultant", "invoice_date": "2025-04-24", "invoice_number": "HRGP012025", "joining_date": "2025-04-14", "ctc": 1600000, "fee_rate": 0.06, "bill_value": 96000, "gst": 17280, "tds": 1920, "status": "Received"},
    {"name": "Soumen Chakraborty", "client": "EY-GGN", "role": "Senior Consultant", "invoice_date": "2025-05-09", "invoice_number": "HRGP022025", "joining_date": "2025-05-02", "ctc": 2020000, "fee_rate": 0.06, "bill_value": 121200, "gst": 21816, "tds": 2424, "status": "Received"},
    {"name": "Pratik Sanjay Nerkar", "client": "EY-GGN", "role": "Senior Consultant", "invoice_date": "2025-07-08", "invoice_number": "HRGP092025", "joining_date": "2025-07-01", "ctc": 2020000, "fee_rate": 0.06, "bill_value": 121200, "gst": 21816, "tds": 2424, "status": "Received"},
    {"name": "Karan Kumar", "client": "H&B", "role": "Design Engineer - Product Design", "invoice_date": "2025-07-21", "invoice_number": "HRGP102025", "joining_date": "2025-07-11", "ctc": 548292, "fee_rate": 0.0833, "bill_value": 45672.72, "gst": 8221.09, "tds": 4567.27, "status": "Received"},
    {"name": "Madhu Kiran Kummari Panyam", "client": "EY-Hyd", "role": "Manager", "invoice_date": "2025-08-08", "invoice_number": "HRGP122025", "joining_date": "2025-08-07", "ctc": 2400000, "fee_rate": 0.06, "bill_value": 144000, "gst": 25920, "tds": 2880, "status": "Received"},
    {"name": "Shashank Singh", "client": "H&B", "role": "Aftermarket sales engineer", "invoice_date": "2025-08-18", "invoice_number": "HRGP132025", "joining_date": "2025-08-11", "ctc": 738571, "fee_rate": 0.0833, "bill_value": 61522.96, "gst": 11074.13, "tds": 6152.30, "status": "Received"},
    {"name": "Mayank Solanki", "client": "Triam", "role": "Director - Container Technology", "invoice_date": "2025-09-11", "invoice_number": "HRGP162025", "joining_date": "2025-09-08", "ctc": 8600000, "fee_rate": 0.0833, "bill_value": 716380, "gst": 128948.40, "tds": 84532.84, "status": "Pending"},
    {"name": "-", "client": "Taggd", "role": "-", "invoice_date": "2025-09-29", "invoice_number": "HRGP172025", "joining_date": "-", "ctc": 0, "fee_rate": 0, "bill_value": 1167743, "gst": 210193, "tds": 116774.30, "status": "Pending"},
]


def read_json_store(path: Path, default):
    if not path.exists():
        return default
    raw = path.read_text(encoding="utf-8").strip()
    if not raw:
        return default
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        backup = path.with_suffix(path.suffix + f".invalid_{datetime.now():%Y%m%d_%H%M%S}.bak")
        path.replace(backup)
        return default


def read_json_list(path: Path) -> list:
    data = read_json_store(path, [])
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and isinstance(data.get("value"), list):
        return data["value"]
    if isinstance(data, dict):
        return [data]
    return []


def read_json_dict(path: Path) -> dict:
    data = read_json_store(path, {})
    return data if isinstance(data, dict) else {}


@dataclass
class Invoice:
    candidate_name: str
    client_alias: str
    role: str
    ctc: float
    joining_date: str
    reference: str
    invoice_number: str
    invoice_date: str
    fee_rate: float
    gst_type: str
    bill_value: float
    cgst: float
    sgst: float
    igst: float
    gst: float
    gross: float
    tds: float
    net_income: float
    amount_tbr: float
    amount_words: str
    status: str = "Pending"
    service_items: list[dict] | None = None


TAGGD_DEFAULT_SERVICE_ITEMS = [
    {"description": "Recruiters Salaries", "count": 41, "rate": 0, "amount": 1416167},
    {"description": "Seats Cost (Recruiters)", "count": 20, "rate": 4500, "amount": 90000},
    {"description": "Manager Salary", "count": 1, "rate": 50000, "amount": 50000},
    {"description": "Seats Cost (Manager)", "count": 1, "rate": 4500, "amount": 4500},
    {"description": "Partial Laptop's (1) Cost Payment - 2/6", "count": 2, "rate": 3333, "amount": 6666},
    {"description": "Partial Laptop's (2) Cost Payment - 3/6", "count": 2, "rate": 3333, "amount": 6666},
    {"description": "Mark-up Charges", "count": 15, "rate": 0, "amount": 234100},
]


def load_mis() -> list[dict]:
    return BASE_MIS + read_json_list(MIS_STORE)


def load_references() -> list[str]:
    saved = read_json_list(REFERENCES_STORE)
    refs = []
    for ref in REFERENCES + saved:
        clean = str(ref).strip()
        if clean and clean.lower() not in {r.lower() for r in refs}:
            refs.append(clean)
    return refs


def load_clients() -> dict:
    clients = {alias: data.copy() for alias, data in CLIENTS.items()}
    saved = read_json_dict(CLIENTS_STORE)
    for alias, data in saved.items():
        clean_alias = str(alias).strip()
        if clean_alias:
            merged = clients.get(clean_alias, {}).copy()
            merged.update(data)
            clients[clean_alias] = merged
    return clients


def hsn_sac_for(client: dict) -> str:
    return str(client.get("hsn_sac") or "998519").strip()


def save_client(form: dict) -> str:
    alias = form.get("new_client_alias", "").strip()
    name = form.get("new_client_name", "").strip()
    if not alias and name:
        alias = re.sub(r"[^A-Za-z0-9]+", "-", name).strip("-")[:24] or name
    if not alias:
        return ""
    saved = read_json_dict(CLIENTS_STORE)
    existing = load_clients()
    saved[alias] = {
        "name": name or existing.get(alias, {}).get("name", alias),
        "address": form.get("new_client_address", "").strip() or existing.get(alias, {}).get("address", ""),
        "gstin": form.get("new_client_gstin", "").replace("GSTIN:", "").replace("GSTIN", "").strip() or existing.get(alias, {}).get("gstin", ""),
        "pan": form.get("new_client_pan", "").replace("PAN:", "").replace("PAN No:", "").strip() or existing.get(alias, {}).get("pan", ""),
        "state": form.get("new_client_state", "").strip() or existing.get(alias, {}).get("state", ""),
        "state_code": existing.get(alias, {}).get("state_code", ""),
        "buyer_po": existing.get(alias, {}).get("buyer_po", ""),
        "hsn_sac": existing.get(alias, {}).get("hsn_sac", ""),
        "spoc": existing.get(alias, {}).get("spoc", ""),
    }
    CLIENTS_STORE.write_text(json.dumps(saved, indent=2), encoding="utf-8")
    return alias


def save_reference(ref: str) -> str:
    clean = ref.strip()
    if not clean:
        return ""
    saved = read_json_list(REFERENCES_STORE)
    if clean.lower() not in {str(r).strip().lower() for r in REFERENCES + saved}:
        saved.append(clean)
        REFERENCES_STORE.write_text(json.dumps(saved, indent=2), encoding="utf-8")
    return clean


def save_generated(row: dict) -> None:
    rows = read_json_list(MIS_STORE)
    rows.append(row)
    MIS_STORE.write_text(json.dumps(rows, indent=2), encoding="utf-8")


def export_mis_excel() -> Path:
    rows = load_mis()
    wb = Workbook()
    ws = wb.active
    ws.title = "MIS"
    headers = [
        "Name", "Company", "Position/Role", "Invoice Date", "Invoice #",
        "Joining Date", "CTC", "%age", "Bill Value (w/o GST)",
        "GST Amount", "Bill Value (incl GST)", "TDS",
        "Net Income(Less TDS/GST)", "Amount TBR (incl GST)",
        "GST Paid", "Status",
    ]
    ws.append(headers)

    for row in rows:
        bill_value = float(row.get("bill_value") or 0)
        gst = float(row.get("gst") or 0)
        gross = float(row.get("gross") or row.get("total") or bill_value + gst)
        tds = float(row.get("tds") or round(bill_value * 0.10, 2))
        net_income = float(row.get("net_income") or round(bill_value - tds, 2))
        amount_tbr = float(row.get("amount_tbr") or round(gross - tds, 2))
        client_alias = row.get("client") or row.get("client_alias") or ""
        company = load_clients().get(client_alias, {}).get("name", client_alias)
        ws.append([
            row.get("name") or row.get("candidate_name") or "",
            company,
            row.get("role") or "",
            row.get("invoice_date") or "",
            row.get("invoice_number") or "",
            row.get("joining_date") or "",
            row.get("ctc") or "",
            row.get("fee_rate") or "",
            bill_value,
            gst,
            gross,
            tds,
            net_income,
            amount_tbr,
            row.get("gst_paid") or "",
            row.get("status") or "Pending",
        ])

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for idx in [7, 9, 10, 11, 12, 13, 14]:
            row[idx - 1].number_format = '#,##0.00'
        row[7].number_format = '0.00%'
    widths = [26, 34, 28, 14, 16, 14, 14, 10, 18, 14, 18, 12, 22, 20, 12, 14]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width
    ws.freeze_panes = "A2"
    wb.save(MIS_EXCEL_PATH)
    return MIS_EXCEL_PATH


def next_invoice_number() -> str:
    numbers = []
    for row in load_mis():
        match = re.search(r"HRGP(\d+)(\d{4})$", row.get("invoice_number", ""))
        if match:
            numbers.append(int(match.group(1)))
    return f"HRGP{max(numbers, default=0) + 1:02d}{date.today().year}"


def fee_rate_for(client_alias: str) -> float:
    for row in reversed(load_mis()):
        if row.get("client") == client_alias and float(row.get("fee_rate") or 0) > 0:
            return float(row["fee_rate"])
    if client_alias.startswith("EY"):
        return 0.06
    return 0.0833


def indian_currency(amount: float) -> str:
    sign = "-" if amount < 0 else ""
    amount = abs(float(amount))
    whole = int(round(amount))
    s = str(whole)
    if len(s) > 3:
        last = s[-3:]
        rest = s[:-3]
        parts = []
        while len(rest) > 2:
            parts.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        s = ",".join(parts + [last])
    return f"{sign}₹ {s}"


ONES = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]


def two_digit_words(n: int) -> str:
    if n < 20:
        return ONES[n]
    return " ".join(x for x in [TENS[n // 10], ONES[n % 10]] if x)


def number_to_words_indian(amount: float) -> str:
    n = int(round(amount))
    if n == 0:
        return "Zero Rupees"
    parts = []
    crore, n = divmod(n, 10000000)
    lakh, n = divmod(n, 100000)
    thousand, n = divmod(n, 1000)
    hundred, n = divmod(n, 100)
    if crore:
        parts.append(two_digit_words(crore) + " Crore")
    if lakh:
        parts.append(two_digit_words(lakh) + " Lakh")
    if thousand:
        parts.append(two_digit_words(thousand) + " Thousand")
    if hundred:
        parts.append(ONES[hundred] + " Hundred")
    if n:
        parts.append(("and " if parts else "") + two_digit_words(n))
    return " ".join(parts) + " Rupees"


def parse_float(value, default: float = 0.0) -> float:
    try:
        return float(str(value or "").replace(",", "").strip())
    except ValueError:
        return default


def taggd_item_kind(description: str) -> str:
    text = description.lower()
    if "mark" in text and "up" in text:
        return "markup"
    if "laptop" in text:
        return "laptop"
    if "seat" in text:
        return "seat"
    if "recruiter" in text and "salar" in text:
        return "recruiter_salary"
    if "manager" in text and "salar" in text:
        return "manager_salary"
    return "other"


def normalize_taggd_service_items(items: list[dict]) -> list[dict]:
    normalized = []
    markup_rows = []
    markup_base = 0.0
    for item in items:
        description = str(item.get("description") or "").strip()
        if not description:
            continue
        kind = taggd_item_kind(description)
        count = parse_float(item.get("count"))
        rate = parse_float(item.get("rate"))
        amount = parse_float(item.get("amount"))
        if kind == "seat":
            rate = 4500
            amount = count * rate
            markup_base += amount
        elif kind == "laptop":
            rate = 3333
            amount = count * rate
        elif kind in {"recruiter_salary", "manager_salary"}:
            if not amount and count and rate:
                amount = count * rate
            markup_base += amount
        elif kind == "markup":
            count = 15
            rate = 0
            markup_rows.append({
                "description": description,
                "count": count,
                "rate": rate,
                "amount": 0,
            })
            continue
        elif not amount and count and rate:
            amount = count * rate
        normalized.append({
            "description": description,
            "count": count,
            "rate": rate,
            "amount": round(amount, 2),
        })
    markup_amount = round(markup_base * 0.15, 2)
    if markup_rows:
        for row in markup_rows:
            row["amount"] = markup_amount
            normalized.append(row)
    else:
        normalized.append({"description": "Mark-up Charges", "count": 15, "rate": 0, "amount": markup_amount})
    return normalized


def parse_taggd_service_items(form: dict) -> list[dict]:
    raw = (form.get("taggd_service_items") or "").strip()
    if raw:
        try:
            items = json.loads(raw)
        except json.JSONDecodeError:
            items = []
    else:
        items = TAGGD_DEFAULT_SERVICE_ITEMS
    cleaned = []
    for item in items:
        description = str(item.get("description") or "").strip()
        if not description:
            continue
        count = parse_float(item.get("count"))
        rate = parse_float(item.get("rate"))
        amount = parse_float(item.get("amount"))
        if not amount and count and rate:
            amount = round(count * rate, 2)
        cleaned.append({
            "description": description,
            "count": count,
            "rate": rate,
            "amount": round(amount, 2),
        })
    return normalize_taggd_service_items(cleaned or TAGGD_DEFAULT_SERVICE_ITEMS[:])


def taggd_service_items_for_form(form: dict) -> list[dict]:
    raw = str(form.get("taggd_service_items") or "").strip()
    if raw:
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            parsed = []
    else:
        parsed = TAGGD_DEFAULT_SERVICE_ITEMS[:]
    return parsed if isinstance(parsed, list) else TAGGD_DEFAULT_SERVICE_ITEMS[:]


def taggd_recruiter_salary_total(form: dict) -> float:
    items = taggd_service_items_for_form(form)
    for item in items:
        desc = str(item.get("description") or "").strip().lower()
        if "recruiter" in desc and "salary" in desc:
            return round(parse_float(item.get("amount")) or 0, 2)
    for item in TAGGD_DEFAULT_SERVICE_ITEMS:
        desc = str(item.get("description") or "").strip().lower()
        if "recruiter" in desc and "salary" in desc:
            return round(parse_float(item.get("amount")) or 0, 2)
    return 0.0


def extract_pdf_text_and_amount(pdf_bytes: bytes) -> dict:
    result = {
        "text": "",
        "amount": None,
        "confidence": "low",
        "status": "manual_review",
        "note": "",
    }
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as exc:
        result["status"] = "invalid_pdf"
        result["note"] = f"Unable to open PDF: {exc}"
        return result

    pages = []
    ocr_errors = []
    for page in doc:
        page_text = (page.get_text("text") or "").strip()
        if not page_text:
            try:
                textpage = page.get_textpage_ocr()
                page_text = (page.get_text("text", textpage=textpage) or "").strip()
            except Exception as exc:
                ocr_errors.append(str(exc))
        if page_text:
            pages.append(page_text)
    doc.close()

    text = "\n".join(pages).strip()
    result["text"] = text
    if not text:
        result["note"] = "No text layer found and OCR is unavailable on this server."
        if ocr_errors:
            result["note"] += " " + ocr_errors[-1]
        return result

    patterns = [
        r"(?:gross(?:\s+salary|\s+pay|\s+earnings)?|net(?:\s+salary|\s+pay)?|take\s*home|salary\s+payable|earnings|amount\s+paid)\s*[:\-]?\s*(?:rs\.?|inr)?\s*([0-9][0-9,]*(?:\.[0-9]{1,2})?)",
        r"(?:^|\n)\s*(?:gross(?:\s+salary|\s+pay)?|net(?:\s+salary|\s+pay)?|salary|earnings)\s+([0-9][0-9,]*(?:\.[0-9]{1,2})?)",
        r"(?:rs\.?|inr)\s*([0-9][0-9,]*(?:\.[0-9]{1,2})?)",
    ]
    candidates = []
    lowered = text.lower()
    for pattern in patterns:
        for match in re.finditer(pattern, lowered, re.IGNORECASE):
            amount_raw = match.group(1)
            try:
                amount = float(amount_raw.replace(",", ""))
            except ValueError:
                continue
            window = lowered[max(0, match.start() - 50): match.end() + 50]
            score = 1
            if "gross" in window:
                score += 4
            if "net" in window:
                score += 3
            if "salary" in window:
                score += 3
            if "take home" in window or "payable" in window:
                score += 2
            candidates.append((score, amount, window))
    if candidates:
        candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
        best = candidates[0]
        result["amount"] = round(best[1], 2)
        result["confidence"] = "medium" if best[0] < 7 else "high"
        result["status"] = "parsed"
        result["note"] = f"Matched on text near: {best[2][:80].strip()}"
    else:
        result["note"] = "Text extracted, but no salary amount matched a known pattern."
    return result

def reconcile_salary_slips(files, target_total: float) -> dict:
    rows = []
    parsed_total = 0.0
    unreadable = 0
    for uploaded in files:
        if not uploaded or not getattr(uploaded, 'filename', ''):
            continue
        pdf_bytes = uploaded.read()
        slip = extract_pdf_text_and_amount(pdf_bytes)
        amount = slip.get("amount")
        if amount is None:
            unreadable += 1
        else:
            parsed_total += float(amount)
        rows.append({
            "filename": uploaded.filename,
            "status": slip.get("status") or "manual_review",
            "amount": amount,
            "confidence": slip.get("confidence") or "low",
            "note": slip.get("note") or "",
        })
    parsed_total = round(parsed_total, 2)
    target_total = round(float(target_total or 0), 2)
    return {
        "rows": rows,
        "parsed_total": parsed_total,
        "target_total": target_total,
        "difference": round(parsed_total - target_total, 2),
        "matched_count": len([r for r in rows if r.get("amount") is not None]),
        "unreadable_count": unreadable,
        "total_files": len(rows),
        "note": "", 
    }


def build_invoice(form: dict) -> Invoice:
    alias = form["client_alias"]
    client = load_clients()[alias]
    service_items = parse_taggd_service_items(form) if alias == "Taggd" else None
    if service_items:
        ctc = sum(float(item.get("amount") or 0) for item in service_items)
        fee_rate = 0
        bill_value = round(ctc, 2)
    else:
        ctc = float(form["ctc"])
        fee_rate = fee_rate_for(alias)
        bill_value = round(ctc * fee_rate, 2)
    same_state = client["state"].lower() == COMPANY["state"].lower()
    if same_state:
        cgst = round(bill_value * 0.09, 2)
        sgst = round(bill_value * 0.09, 2)
        igst = 0
        gst_type = "CGST + SGST"
    else:
        cgst = sgst = 0
        igst = round(bill_value * 0.18, 2)
        gst_type = "IGST"
    gst = cgst + sgst + igst
    gross = round(bill_value + gst, 2)
    tds = 0 if service_items else round(bill_value * 0.10, 2)
    net_income = round(bill_value - tds, 2)
    amount_tbr = round(gross - tds, 2)
    return Invoice(
        candidate_name=(form.get("candidate_name") or "").strip() or ("Taggd" if service_items else ""),
        client_alias=alias,
        role=(form.get("role") or "").strip() or ("Monthly Service Invoice" if service_items else ""),
        ctc=ctc,
        joining_date=form.get("joining_date") or date.today().isoformat(),
        reference=form.get("reference") or client.get("spoc", ""),
        invoice_number=form.get("invoice_number") or next_invoice_number(),
        invoice_date=date.today().isoformat(),
        fee_rate=fee_rate,
        gst_type=gst_type,
        bill_value=bill_value,
        cgst=cgst,
        sgst=sgst,
        igst=igst,
        gst=gst,
        gross=gross,
        tds=tds,
        net_income=net_income,
        amount_tbr=amount_tbr,
        amount_words=number_to_words_indian(gross),
        service_items=service_items,
    )


class TinyPDF:
    def __init__(self) -> None:
        self.ops = []
        self.images = []

    def color(self, r: float, g: float, b: float, stroke: bool = False) -> None:
        op = "RG" if stroke else "rg"
        self.ops.append(f"{r:.4f} {g:.4f} {b:.4f} {op}")

    def text(self, x: int, y: int, value: str, size: float = 10, bold: bool = False, color: tuple[float, float, float] | None = None, align: str = "left", width: int = 0) -> None:
        self.color(*(color or (0, 0, 0)))
        value = value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        if align != "left" and width:
            rough = len(value) * size * 0.48
            if align == "right":
                x = x + width - rough
            elif align == "center":
                x = x + (width - rough) / 2
        font = "F2" if bold else "F1"
        self.ops.append(f"BT /{font} {size} Tf {x:.2f} {y:.2f} Td ({value}) Tj ET")

    def wrapped(self, x: int, y: int, value: str, chars: int, size: float = 8, leading: int = 12, bold: bool = False) -> int:
        for line in wrap(value, chars) or [""]:
            self.text(x, y, line, size=size, bold=bold)
            y -= leading
        return y

    def line(self, x1: int, y1: int, x2: int, y2: int) -> None:
        self.ops.append(f"{x1} {y1} m {x2} {y2} l S")

    def rect(self, x: int, y: int, w: int, h: int) -> None:
        self.ops.append(f"{x} {y} {w} {h} re S")

    def fill_rect(self, x: int, y: int, w: int, h: int, color: tuple[float, float, float]) -> None:
        self.color(*color)
        self.ops.append(f"{x} {y} {w} {h} re f")

    def image(self, path: Path, x: int, y: int, w: int, h: int) -> None:
        if not path.exists():
            return
        name = f"Im{len(self.images) + 1}"
        if path.suffix.lower() in {".jpg", ".jpeg"}:
            width, height = jpeg_size(path)
            self.images.append((name, width, height, path.read_bytes(), "DCTDecode"))
        else:
            width, height, data = png_rgb(path)
            self.images.append((name, width, height, zlib.compress(data), "FlateDecode"))
        self.ops.append(f"q {w} 0 0 {h} {x} {y} cm /{name} Do Q")

    def save(self, path: Path) -> None:
        stream = "0.3 w\n" + "\n".join(self.ops)
        xobjects = " ".join(f"/{name} {6+i} 0 R" for i, (name, *_rest) in enumerate(self.images))
        objects = [
            "<< /Type /Catalog /Pages 2 0 R >>",
            "<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R /F2 5 0 R >> /XObject << {xobjects} >> >> /Contents {6+len(self.images)} 0 R >>",
            "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            "<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
        ]
        for _name, width, height, image_data, pdf_filter in self.images:
            objects.append(f"<< /Type /XObject /Subtype /Image /Width {width} /Height {height} /ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /{pdf_filter} /Length {len(image_data)} >>\nstream\n{image_data.decode('latin-1')}\nendstream")
        objects.append(
            f"<< /Length {len(stream.encode('latin-1', 'replace'))} >>\nstream\n{stream}\nendstream",
        )
        out = ["%PDF-1.4\n"]
        offsets = [0]
        for i, obj in enumerate(objects, 1):
            offsets.append(sum(len(x.encode("latin-1", "replace")) for x in out))
            out.append(f"{i} 0 obj\n{obj}\nendobj\n")
        xref = sum(len(x.encode("latin-1", "replace")) for x in out)
        out.append(f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n")
        for off in offsets[1:]:
            out.append(f"{off:010d} 00000 n \n")
        out.append(f"trailer << /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF")
        path.write_bytes("".join(out).encode("latin-1", "replace"))


def png_rgb(path: Path) -> tuple[int, int, bytes]:
    data = path.read_bytes()
    if data[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError("Logo must be a PNG file")
    pos = 8
    width = height = color_type = bit_depth = None
    chunks = []
    while pos < len(data):
        length = struct.unpack(">I", data[pos:pos + 4])[0]
        kind = data[pos + 4:pos + 8]
        body = data[pos + 8:pos + 8 + length]
        pos += 12 + length
        if kind == b"IHDR":
            width, height, bit_depth, color_type = struct.unpack(">IIBB", body[:10])
        elif kind == b"IDAT":
            chunks.append(body)
        elif kind == b"IEND":
            break
    if bit_depth != 8 or color_type not in (2, 6):
        raise ValueError("Only 8-bit RGB/RGBA PNG logos are supported")
    raw = zlib.decompress(b"".join(chunks))
    channels = 4 if color_type == 6 else 3
    stride = width * channels
    rows = []
    prev = [0] * stride
    i = 0
    for _ in range(height):
        f = raw[i]
        i += 1
        row = list(raw[i:i + stride])
        i += stride
        for x in range(stride):
            left = row[x - channels] if x >= channels else 0
            up = prev[x]
            up_left = prev[x - channels] if x >= channels else 0
            if f == 1:
                row[x] = (row[x] + left) & 255
            elif f == 2:
                row[x] = (row[x] + up) & 255
            elif f == 3:
                row[x] = (row[x] + ((left + up) // 2)) & 255
            elif f == 4:
                p = left + up - up_left
                pa, pb, pc = abs(p - left), abs(p - up), abs(p - up_left)
                row[x] = (row[x] + (left if pa <= pb and pa <= pc else up if pb <= pc else up_left)) & 255
        rgb = bytearray()
        for x in range(0, stride, channels):
            if channels == 4:
                a = row[x + 3] / 255
                rgb.extend(int(row[x + c] * a + 255 * (1 - a)) for c in range(3))
            else:
                rgb.extend(row[x:x + 3])
        rows.append(bytes(rgb))
        prev = row
    return width, height, b"".join(rows)


def jpeg_size(path: Path) -> tuple[int, int]:
    data = path.read_bytes()
    i = 2
    while i < len(data):
        if data[i] != 0xFF:
            i += 1
            continue
        marker = data[i + 1]
        i += 2
        if marker in (0xD8, 0xD9):
            continue
        length = int.from_bytes(data[i:i + 2], "big")
        if marker in range(0xC0, 0xC4):
            height = int.from_bytes(data[i + 3:i + 5], "big")
            width = int.from_bytes(data[i + 5:i + 7], "big")
            return width, height
        i += length
    raise ValueError("Could not read JPEG dimensions")


def pdf_amount(amount: float) -> str:
    return indian_currency(amount).replace("₹ ", "")


def pdf_for(invoice: Invoice) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    client = load_clients()[invoice.client_alias]
    safe_candidate = re.sub(r"[^A-Za-z0-9._-]+", "_", invoice.candidate_name).strip("_") or "invoice"
    path = OUTPUT_DIR / f"{invoice.invoice_number}_{safe_candidate}.pdf"
    pdf = TinyPDF()
    navy = (0.1216, 0.2196, 0.3922)
    blue = (0.1804, 0.3294, 0.5882)
    pale = (0.9333, 0.9529, 0.9843)
    stripe = (0.9686, 0.9765, 1.0)
    grey = (0.7333, 0.7333, 0.7333)

    pdf.image(PDF_LOGO_PATH if PDF_LOGO_PATH.exists() else LOGO_PATH, 46, 753, 62, 54)
    meta_x, meta_value_x, meta_y = 346, 452, 790
    meta = [
        ("Invoice Number", invoice.invoice_number),
        ("Invoice Date", datetime.fromisoformat(invoice.invoice_date).strftime("%B %d, %Y")),
    ]
    for label, value in meta:
        pdf.text(meta_x, meta_y, label, 10.5, bold=True)
        pdf.text(meta_value_x, meta_y, value, 10.5, bold=True)
        meta_y -= 18
    pdf.text(meta_x, meta_y, "Mode of Payment", 10.5, bold=True)
    pdf.text(meta_value_x, meta_y, "Bank Transfer", 10.5, bold=True)
    pdf.color(*navy, stroke=True)
    pdf.line(46, 739, 549, 739)

    pdf.fill_rect(40, 711, 516, 23, blue)
    pdf.text(40, 718, "TAX INVOICE", 11, bold=True, color=(1, 1, 1), align="center", width=516)

    pdf.fill_rect(40, 555, 258, 151, pale)
    pdf.fill_rect(298, 555, 258, 151, pale)
    for x in (40, 298):
        pdf.color(*grey, stroke=True)
        pdf.rect(x, 555, 258, 151)
    pdf.text(47, 690, "Biller", 11, bold=True, color=navy)
    pdf.text(47, 675, COMPANY["name"], 9, bold=True)
    y = 662
    for line in [COMPANY["address"], f"Contact No: +91 {COMPANY['contact'].replace('+91 ', '')}", f"GSTIN: {COMPANY['gstin']}", f"PAN: {COMPANY['pan']}", f"State Code: {COMPANY['state_code']}", f"HSN/SAC Code: {hsn_sac_for(client)}"]:
        y = pdf.wrapped(47, y, line, 34, size=8, leading=12)
    pdf.text(305, 690, "Billed To", 11, bold=True, color=navy)
    pdf.text(305, 675, client["name"], 9, bold=True)
    y = pdf.wrapped(305, 662, client["address"], 34, size=8, leading=12)
    billed_to_lines = [f"GSTIN: {client['gstin']}", f"PAN: {client['pan']}"]
    if client.get("buyer_po"):
        billed_to_lines.append(f"Buyer PO Number: {client['buyer_po']}")
    billed_to_lines.extend([
        f"Buyer's Spoc: {client.get('spoc', invoice.reference)}",
    ])
    for line in billed_to_lines:
        y = pdf.wrapped(305, y, line, 34, size=8, leading=12)

    if invoice.service_items:
        x0, top, w, row_h = 40, 535, 516, 30
        total_rows = 1 + len(invoice.service_items) + 3
        bottom = top - row_h * total_rows
        pdf.fill_rect(x0, top - row_h, w, row_h, blue)
        cols = [x0, x0 + 258, x0 + 326, x0 + 416, x0 + w]
        for idx in range(total_rows + 1):
            yline = top - row_h * idx
            pdf.color(*grey, stroke=True)
            pdf.line(x0, yline, x0 + w, yline)
            if idx > 1 and idx % 2 == 0:
                pdf.fill_rect(x0, yline, w, row_h, stripe)
        for x in cols:
            pdf.line(x, bottom, x, top)
        headers = [("Description of Services", x0 + 5, top - 15, 0), ("Count", x0 + 258, top - 15, 68), ("Rate", x0 + 326, top - 15, 90), ("Amount", x0 + 416, top - 15, 100)]
        for h, x, y, cw in headers:
            pdf.text(x, y, h, 8.5, bold=True, color=(1, 1, 1), align="center" if cw else "left", width=cw)
        row_y = top - 36
        for item in invoice.service_items:
            pdf.wrapped(x0 + 5, row_y + 10, item["description"], 34, size=8, leading=11)
            pdf.text(x0 + 258, row_y + 6, f"{item['count']:g}", 8.5, align="center", width=68)
            pdf.text(x0 + 326, row_y + 6, pdf_amount(item["rate"]) if item["rate"] else "", 8.5, align="right", width=80)
            pdf.text(x0 + 416, row_y + 6, pdf_amount(item["amount"]), 8.5, align="right", width=90)
            row_y -= row_h
        tax_rows = [("CGST", "9%" if invoice.cgst else "0%", invoice.cgst), ("SGST", "9%" if invoice.sgst else "0%", invoice.sgst), ("Total", "", invoice.gross)]
        for label, rate, amount in tax_rows:
            pdf.text(x0 + 258, row_y + 6, label, 8.5 if label != "Total" else 9, bold=True, color=navy if label == "Total" else None, align="center", width=68)
            pdf.text(x0 + 326, row_y + 6, rate, 8.5, align="center", width=90)
            pdf.text(x0 + 416, row_y + 6, pdf_amount(amount), 8.5 if label != "Total" else 9, bold=label == "Total", color=navy if label == "Total" else None, align="right", width=90)
            row_y -= row_h
        words_y = bottom - 29
        bank_box_y = 118
    else:
        x0, y0, w = 40, 375, 516
        pdf.fill_rect(x0, y0 + 160, w, 21, blue)
        pdf.fill_rect(x0, y0 + 84, w, 76, (1, 1, 1))
        pdf.fill_rect(x0, y0 + 63, w, 21, stripe)
        pdf.fill_rect(x0, y0 + 42, w, 21, (1, 1, 1))
        pdf.fill_rect(x0, y0 + 21, w, 21, stripe)
        pdf.fill_rect(x0, y0, w, 21, pale)
        cols = [x0, x0 + 268, x0 + 361, x0 + 433, x0 + w]
        for x in cols:
            pdf.color(*grey, stroke=True)
            pdf.line(x, y0, x, y0 + 181)
        for yline in [y0, y0 + 21, y0 + 42, y0 + 63, y0 + 84, y0 + 160, y0 + 181]:
            pdf.line(x0, yline, x0 + w, yline)
        headers = [("Description of Services", x0 + 5, y0 + 165, 0), ("HSN/SAC", x0 + 268, y0 + 165, 93), ("Rate", x0 + 361, y0 + 165, 72), ("Amount", x0 + 433, y0 + 165, 83)]
        for h, x, y, cw in headers:
            pdf.text(x, y, h, 8.5, bold=True, color=(1, 1, 1), align="center" if cw else "left", width=cw)
        doj = datetime.fromisoformat(invoice.joining_date).strftime("%B %d, %Y")
        desc_y = y0 + 143
        pdf.text(x0 + 5, desc_y, "Recruitment / Placement Services", 10, bold=True)
        pdf.text(x0 + 5, desc_y - 15, f"Candidate: {invoice.candidate_name}", 9.5, bold=True)
        pdf.text(x0 + 5, desc_y - 30, f"D.O.J: {doj}", 9)
        pdf.wrapped(x0 + 5, desc_y - 45, f"Role: {invoice.role.title()}", 36, size=9, leading=13)
        pdf.text(x0 + 268, y0 + 120, hsn_sac_for(client), 8.5, align="center", width=93)
        pdf.text(x0 + 361, y0 + 120, f"{invoice.fee_rate * 100:.2f}%", 8.5, align="center", width=72)
        pdf.text(x0 + 433, y0 + 120, pdf_amount(invoice.bill_value), 8.5, align="right", width=74)
        pdf.text(x0 + 5, y0 + 69, "Offered CTC", 8.5, bold=True)
        pdf.text(x0 + 433, y0 + 69, pdf_amount(invoice.ctc), 8.5, align="right", width=74)
        gst_rows = [("CGST", "9%" if invoice.cgst else "0%", invoice.cgst), ("IGST", "18%" if invoice.igst else "0%", invoice.igst), ("Total", "", invoice.gross)]
        for row_y, (label, rate, amount) in zip([y0 + 46, y0 + 25, y0 + 4], gst_rows):
            pdf.text(x0 + 268, row_y, label, 8.5 if label != "Total" else 9, bold=True, color=navy if label == "Total" else None, align="center", width=93)
            pdf.text(x0 + 361, row_y, rate, 8.5, align="center", width=72)
            pdf.text(x0 + 433, row_y, pdf_amount(amount), 8.5 if label != "Total" else 9, bold=label == "Total", color=navy if label == "Total" else None, align="right", width=74)
        words_y = 346
        bank_box_y = 235

    pdf.fill_rect(40, words_y, 516, 23, pale)
    pdf.color(*grey, stroke=True)
    pdf.rect(40, words_y, 516, 23)
    pdf.text(46, words_y + 8, f"Amount in words: {invoice.amount_words.replace(' Rupees', '')} Only", 8.5, bold=True, color=navy)

    pdf.rect(40, bank_box_y, 516, 103)
    pdf.line(324, bank_box_y, 324, bank_box_y + 103)
    pdf.text(47, bank_box_y + 84, "Declaration:", 8.5, bold=True)
    pdf.wrapped(47, bank_box_y + 58, "We declare that this invoice shows the actual price of the services described and that all particulars are true and correct.", 62, size=8, leading=13)
    pdf.text(331, bank_box_y + 84, "Company's Bank Details", 8.5, bold=True)
    bank_lines = [
        f"Bank Name: {COMPANY['bank']}",
        f"A/c No: {COMPANY['account']}",
        "Branch And IFSC :",
        f"{COMPANY['branch']}",
        f"IFSC : {COMPANY['ifsc']}",
    ]
    y = bank_box_y + 71
    for i, line in enumerate(bank_lines):
        y = pdf.wrapped(331, y, line, 45, size=8, leading=13, bold=i == 2)
    pdf.text(418, bank_box_y - 47, "(Authorized Signatory)", 8.5, bold=True)
    pdf.color(*navy, stroke=True)
    pdf.line(46, bank_box_y - 63, 549, bank_box_y - 63)
    pdf.text(46, bank_box_y - 79, f"{COMPANY['name']}| GSTIN: {COMPANY['gstin']} | Tower 8/1202, Orchid Petals, Sector 49, Gurgaon, 122018, Haryana, India", 7, color=(0.3333, 0.3333, 0.3333), align="center", width=503)
    pdf.save(path)
    return path


HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>HR Guru Invoice Generator</title>
  <style>
    :root { --navy:#071b3a; --navy2:#0c2d5b; --line:#d8e2ef; --ink:#102033; --muted:#637083; }
    * { box-sizing: border-box; }
    body { margin:0; font-family: Inter, Arial, sans-serif; color:var(--ink); background:#f4f7fb; }
    header { background:var(--navy); color:#fff; padding:18px 32px; display:flex; align-items:center; gap:16px; }
    header img { width:52px; height:52px; background:#fff; border-radius:8px; object-fit:contain; padding:4px; }
    main { max-width:1120px; margin:26px auto; padding:0 18px; display:grid; grid-template-columns: 420px 1fr; gap:22px; }
    section { background:#fff; border:1px solid var(--line); border-radius:8px; box-shadow:0 8px 24px rgba(7,27,58,.07); }
    .panel { padding:22px; }
    h1 { font-size:24px; margin:0; letter-spacing:0; }
    h2 { font-size:17px; margin:0 0 18px; color:var(--navy); }
    label { display:block; font-size:13px; color:var(--muted); margin:14px 0 6px; }
    input, select { width:100%; min-height:42px; border:1px solid #cfd9e8; border-radius:6px; padding:9px 11px; font-size:15px; }
    button, .button { border:0; border-radius:6px; padding:11px 15px; background:var(--navy); color:#fff; font-weight:700; cursor:pointer; text-decoration:none; display:inline-block; }
    .button.secondary { background:#e7edf6; color:var(--navy); }
    .actions { display:flex; gap:10px; margin-top:18px; flex-wrap:wrap; }
    .mode-grid { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:16px; margin-bottom:20px; }
    .mode-card { border:1px solid var(--line); border-radius:10px; padding:18px; background:#fff; box-shadow:0 8px 24px rgba(7,27,58,.05); cursor:pointer; text-align:left; }
    .mode-card:hover { border-color:var(--navy2); box-shadow:0 10px 30px rgba(7,27,58,.08); }
    .mode-card.active { border-color:var(--navy2); background:#f5f9ff; }
    .mode-card strong { display:block; color:var(--navy); font-size:18px; margin-bottom:6px; }
    .mode-card span { display:block; color:var(--muted); font-size:13px; line-height:1.45; }
    .invoice-form-wrap.hidden { display:none; }
    .recon-card { margin-top:18px; border:1px solid var(--line); border-radius:10px; padding:18px; background:#f8fbff; }
    .recon-head { display:flex; justify-content:space-between; gap:12px; align-items:flex-start; margin-bottom:10px; }
    .recon-head strong { color:var(--navy); font-size:16px; }
    .recon-grid { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:12px; margin-top:14px; }
    .recon-summary { margin-top:14px; border-top:1px solid var(--line); padding-top:14px; }
    .recon-summary-grid { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; }
    .recon-stat { background:#fff; border:1px solid var(--line); border-radius:8px; padding:10px 12px; }
    .recon-stat span { display:block; color:var(--muted); font-size:11px; text-transform:uppercase; font-weight:700; margin-bottom:4px; }
    .recon-stat strong { color:var(--navy); font-size:16px; }
    .recon-table-wrap { margin-top:12px; overflow:auto; border:1px solid var(--line); border-radius:8px; background:#fff; }
    .recon-table { width:100%; border-collapse:collapse; margin:0; }
    .recon-table th, .recon-table td { border-bottom:1px solid var(--line); padding:10px 12px; vertical-align:top; font-size:13px; }
    .recon-table th { background:#eff4fb; color:var(--navy); }
    .recon-status { display:inline-flex; align-items:center; padding:3px 8px; border-radius:999px; font-size:11px; font-weight:800; background:#e7edf6; color:var(--navy); }
    .recon-status.good { background:#e3f7ec; color:#137a3a; }
    .recon-status.warn { background:#fff3d9; color:#a86c00; }
    .recon-status.bad { background:#ffe2e2; color:#a22020; }
    .invoice { padding:28px; }
    .invoice-head { display:flex; justify-content:space-between; align-items:start; gap:20px; border-bottom:2px solid var(--navy); padding-bottom:16px; }
    .brand { display:flex; gap:14px; align-items:center; }
    .brand img { width:72px; height:62px; object-fit:contain; }
    .badge { color:#fff; background:var(--navy2); padding:6px 10px; border-radius:4px; font-size:12px; display:inline-block; }
    .grid { display:grid; grid-template-columns:1fr 1fr; gap:18px; margin:22px 0; }
    .box { border:1px solid var(--line); border-radius:8px; padding:14px; min-height:138px; }
    .box strong { color:var(--navy); display:block; margin-bottom:6px; }
    .candidate-field.hidden-for-taggd { display:none; }
    table { width:100%; border-collapse:collapse; margin:18px 0; }
    th { background:var(--navy); color:#fff; text-align:left; padding:10px; font-size:13px; }
    td { border:1px solid var(--line); padding:10px; vertical-align:top; }
    td.num, th.num { text-align:right; }
    .totals { margin-left:auto; width:360px; }
    .totals td:first-child { font-weight:700; color:var(--navy); }
    .words { background:#f0f5fb; border-left:4px solid var(--navy); padding:12px; margin-top:12px; }
    .muted { color:var(--muted); font-size:13px; line-height:1.45; }
    .check-row { display:flex; align-items:center; gap:10px; margin-top:16px; font-weight:700; color:var(--navy); }
    .check-row input { width:18px; min-height:18px; }
    dialog { width:min(560px, calc(100vw - 28px)); border:0; border-radius:8px; padding:0; box-shadow:0 24px 70px rgba(7,27,58,.25); }
    dialog.wide { width:min(900px, calc(100vw - 28px)); }
    dialog::backdrop { background:rgba(7,27,58,.42); }
    .modal-head { background:var(--navy); color:#fff; padding:15px 18px; font-weight:800; }
    .modal-body { padding:18px; }
    .modal-actions { display:flex; justify-content:flex-end; gap:10px; padding:0 18px 18px; }
    .taggd-row { display:grid; grid-template-columns: minmax(220px,1fr) 80px 120px 130px; gap:10px; align-items:end; margin-bottom:10px; }
    .taggd-row label { margin-top:0; }
    .taggd-total { text-align:right; font-weight:800; color:var(--navy); margin-top:12px; }
    @media (max-width: 900px) { main { grid-template-columns:1fr; } .grid { grid-template-columns:1fr; } .totals { width:100%; } }
  </style>
</head>
<body>
<header>
  <img src="{{ url_for(route('logo')) }}" alt="HR Guru logo">
  <div><h1>Invoice Generator</h1><div>HR Guru Placement Services Pvt Ltd</div></div>
</header>
<main>
  <section class="panel">
    <h2>Select Invoice Type</h2>
    <div class="mode-grid">
      <button type="button" class="mode-card {% if selected_mode == 'taggd' %}active{% endif %}" data-mode="taggd">
        <strong>Taggd</strong>
        <span>Open the Taggd invoice flow. Company name is fixed to Taggd and the service markup rows are handled automatically.</span>
      </button>
      <button type="button" class="mode-card {% if selected_mode == 'non_taggd' %}active{% endif %}" data-mode="non_taggd">
        <strong>Non Taggd</strong>
        <span>Open the standard invoice flow for every client except Taggd.</span>
      </button>
    </div>
    <div id="invoiceFormWrap" class="invoice-form-wrap {% if not selected_mode %}hidden{% endif %}">
      <h2>Candidate & Invoice Details</h2>
      <form method="post" action="{{ url_for(route('generate')) }}">
        <input type="hidden" name="invoice_mode" id="invoiceModeInput" value="{{ selected_mode or '' }}">
        <div class="candidate-field"><label>Candidate Name</label><input name="candidate_name" required value="{{ form.candidate_name or '' }}"></div>
        <label>Company Name</label><select id="clientAlias" name="client_alias" required>{% for alias, c in clients.items() %}<option value="{{ alias }}" {% if form.client_alias == alias %}selected{% endif %}>{{ alias }} - {{ c.name }}</option>{% endfor %}</select>
        <label class="check-row"><input id="newCompanyToggle" type="checkbox"> New Company</label>
        <input type="hidden" name="new_client_alias">
        <input type="hidden" name="new_client_name">
        <input type="hidden" name="new_client_address">
        <input type="hidden" name="new_client_gstin">
        <input type="hidden" name="new_client_pan">
        <input type="hidden" name="new_client_state">
        <input type="hidden" name="taggd_service_items" id="taggdServiceItems">
        <div class="candidate-field"><label>CTC Amount</label><input name="ctc" type="number" step="1" min="0" required value="{{ form.ctc or '' }}"></div>
        <div class="candidate-field"><label>Role</label><input name="role" required value="{{ form.role or '' }}"></div>
        <div class="candidate-field"><label>Date of Joining</label><input name="joining_date" type="date" required value="{{ form.joining_date or '' }}"></div>
        <div class="candidate-field"><label>Reference Name</label><select name="reference">{% for ref in references %}<option {% if form.reference == ref %}selected{% endif %}>{{ ref }}</option>{% endfor %}</select></div>
        <div class="candidate-field"><label>Add New Reference</label><input name="new_reference" placeholder="Type a new HR reference name"></div>
        <label>Invoice Number</label><input name="invoice_number" required value="{{ form.invoice_number or next_invoice }}">
        <div class="actions"><button type="submit">Generate Invoice</button></div>
      </form>
      {% if selected_mode == 'taggd' %}
      <div class="recon-card">
        <div class="recon-head">
          <div>
            <strong>Salary Slip Reconciliation</strong>
            <div class="muted">Upload scanned salary slip PDFs. The app will try to read text from each file and compare the summed salary amount with the Taggd recruiter salary total.</div>
          </div>
          <div class="badge">Target: {{ money(reconcile_target or 0) }}</div>
        </div>
        <form method="post" action="{{ url_for(route('reconcile_salary_slips_route')) }}" enctype="multipart/form-data" id="salaryReconForm">
          <input type="hidden" name="invoice_mode" value="taggd">
          <input type="hidden" name="taggd_service_items" id="reconcileTaggdServiceItems" value="{{ form.get('taggd_service_items') or '' }}">
          <label>Salary Slip PDFs</label>
          <input type="file" name="salary_slips" accept="application/pdf" multiple required>
          <div class="actions"><button type="submit" class="button secondary">Run Reconciliation</button></div>
        </form>
        {% if reconcile_result %}
        <div class="recon-summary">
          <div class="recon-summary-grid">
            <div class="recon-stat"><span>Files</span><strong>{{ reconcile_result.total_files }}</strong></div>
            <div class="recon-stat"><span>Parsed</span><strong>{{ reconcile_result.matched_count }}</strong></div>
            <div class="recon-stat"><span>Unreadable</span><strong>{{ reconcile_result.unreadable_count }}</strong></div>
            <div class="recon-stat"><span>Difference</span><strong>{{ money(reconcile_result.difference) }}</strong></div>
          </div>
          <div class="recon-table-wrap">
            <table class="recon-table">
              <thead>
                <tr><th>File</th><th>Status</th><th class="num">Amount</th><th>Note</th></tr>
              </thead>
              <tbody>
                {% for row in reconcile_result.rows %}
                <tr>
                  <td>{{ row.filename }}</td>
                  <td><span class="recon-status {% if row.status == 'parsed' %}good{% elif row.status in ['manual_review', 'invalid_pdf'] %}warn{% else %}bad{% endif %}">{{ row.status|replace('_', ' ')|title }}</span></td>
                  <td class="num">{{ money(row.amount or 0) if row.amount is not none else '-' }}</td>
                  <td class="muted">{{ row.note }}</td>
                </tr>
                {% endfor %}
              </tbody>
            </table>
          </div>
          {% if reconcile_result.note %}
          <div class="muted" style="margin-top:10px">{{ reconcile_result.note }}</div>
          {% endif %}
        </div>
        {% endif %}
      </div>
      {% endif %}
    </div>
  </section>
  <section class="invoice">
    {% if invoice %}
      {% set client = clients[invoice.client_alias] %}
      <div class="invoice-head">
        <div class="brand"><img src="{{ url_for(route('logo')) }}"><div><h2>TAX INVOICE</h2><strong>{{ company.name }}</strong><div class="muted">{{ company.address }}</div></div></div>
        <div><span class="badge">{{ invoice.invoice_number }}</span><div class="muted" style="margin-top:8px;">Invoice Date: {{ invoice.invoice_date }}</div></div>
      </div>
      <div class="grid">
        <div class="box"><strong>From</strong>{{ company.name }}<br>{{ company.address }}<br>GSTIN: {{ company.gstin }}<br>PAN: {{ company.pan }}<br>State Code: {{ company.state_code }}<br>HSN/SAC Code: {{ client.hsn_sac or '998519' }}</div>
        <div class="box"><strong>To</strong>{{ client.name }}<br>{{ client.address }}<br>GSTIN: {{ client.gstin }}<br>PAN: {{ client.pan }}{% if client.buyer_po %}<br>Buyer PO Number: {{ client.buyer_po }}{% endif %}<br>Buyer's Spoc: {{ client.spoc or invoice.reference }}<br>GST Type: {{ invoice.gst_type }}</div>
      </div>
      {% if invoice.service_items %}
        <table>
          <thead><tr><th>Description of Services</th><th class="num">Count</th><th class="num">Rate</th><th class="num">Amount</th></tr></thead>
          <tbody>{% for item in invoice.service_items %}<tr><td>{{ item.description }}</td><td class="num">{{ '%g'|format(item.count) }}</td><td class="num">{{ money(item.rate) if item.rate else '' }}</td><td class="num">{{ money(item.amount) }}</td></tr>{% endfor %}</tbody>
        </table>
      {% else %}
        <table>
          <thead><tr><th>Description of Services</th><th>HSN/SAC</th><th class="num">Rate</th><th class="num">Amount</th></tr></thead>
          <tbody><tr><td>{{ invoice.candidate_name }} / D.O.J: {{ invoice.joining_date }}<br>Role: {{ invoice.role.title() }}<br>Reference: {{ invoice.reference }}</td><td>{{ client.hsn_sac or '998519' }}</td><td class="num">{{ '%.2f'|format(invoice.fee_rate * 100) }}%</td><td class="num">{{ money(invoice.bill_value) }}</td></tr></tbody>
        </table>
      {% endif %}
      <table class="totals">
        <tr><td>Bill Value</td><td class="num">{{ money(invoice.bill_value) }}</td></tr>
        <tr><td>CGST</td><td class="num">{{ money(invoice.cgst) }}</td></tr>
        <tr><td>SGST</td><td class="num">{{ money(invoice.sgst) }}</td></tr>
        <tr><td>IGST</td><td class="num">{{ money(invoice.igst) }}</td></tr>
        <tr><td>Total</td><td class="num">{{ money(invoice.gross) }}</td></tr>
        <tr><td>TDS</td><td class="num">{{ money(invoice.tds) }}</td></tr>
        <tr><td>Amount TBR</td><td class="num">{{ money(invoice.amount_tbr) }}</td></tr>
      </table>
      <div class="words">Amount in words: {{ invoice.amount_words }} Only</div>
      <div class="actions"><a class="button" href="{{ url_for(route('download_pdf'), invoice_number=invoice.invoice_number) }}">Download PDF</a><a class="button secondary" href="{{ url_for(route('export_mis')) }}">Export MIS to Excel</a><a class="button secondary" href="{{ url_for(route('mis')) }}">Download MIS JSON</a></div>
    {% else %}
      <h2>Preview</h2>
      <p class="muted">Fill the form and generate the invoice. The app will detect GST type, fee rate, invoice number, totals, amount in words, and append the MIS row with Pending status.</p>
    {% endif %}
  </section>
</main>
<dialog id="companyModal">
  <div class="modal-head">Add New Company</div>
  <div class="modal-body">
    <label>Company Alias</label><input id="modal_client_alias" placeholder="Example: SCL">
    <label>Company Name</label><input id="modal_client_name" placeholder="Full legal company name">
    <label>Company Address</label><input id="modal_client_address" placeholder="Billing address">
    <label>Company GSTIN</label><input id="modal_client_gstin" placeholder="Example: 24ABCDE1234F1Z5">
    <label>Company PAN</label><input id="modal_client_pan" placeholder="Example: ABCDE1234F">
    <label>Company State</label><input id="modal_client_state" placeholder="Example: Gujarat">
  </div>
  <div class="modal-actions">
    <button type="button" class="button secondary" id="cancelCompany">Cancel</button>
    <button type="button" id="saveCompany">Use Company</button>
  </div>
</dialog>
<dialog id="taggdModal" class="wide">
  <div class="modal-head">Taggd Service Details</div>
  <div class="modal-body">
    <div id="taggdRows"></div>
    <div class="actions"><button type="button" class="button secondary" id="addTaggdRow">Add Row</button></div>
    <div class="taggd-total" id="taggdSubtotal">Subtotal: ₹ 0</div>
  </div>
  <div class="modal-actions">
    <button type="button" class="button secondary" id="cancelTaggd">Cancel</button>
    <button type="button" id="saveTaggd">Use Service Details</button>
  </div>
</dialog>
<script>
  const modal = document.getElementById('companyModal');
  const toggle = document.getElementById('newCompanyToggle');
  const clientAlias = document.getElementById('clientAlias');
  const invoiceModeInput = document.getElementById('invoiceModeInput');
  const invoiceFormWrap = document.getElementById('invoiceFormWrap');
  const modeCards = [...document.querySelectorAll('.mode-card')];
  const salaryReconForm = document.getElementById('salaryReconForm');
  const taggdServiceItemsInput = document.getElementById('taggdServiceItems');
  const reconcileTaggdServiceItems = document.getElementById('reconcileTaggdServiceItems');
  const taggdModal = document.getElementById('taggdModal');
  const taggdRows = document.getElementById('taggdRows');
  const taggdDefaults = {{ taggd_defaults|tojson }};
  const map = {
    new_client_alias: 'modal_client_alias',
    new_client_name: 'modal_client_name',
    new_client_address: 'modal_client_address',
    new_client_gstin: 'modal_client_gstin',
    new_client_pan: 'modal_client_pan',
    new_client_state: 'modal_client_state'
  };
  toggle?.addEventListener('change', () => {
    if (toggle.checked) modal.showModal();
  });
  function applyInvoiceMode(mode) {
    const normalized = mode === 'taggd' ? 'taggd' : 'non_taggd';
    if (invoiceModeInput) invoiceModeInput.value = normalized;
    if (invoiceFormWrap) invoiceFormWrap.classList.remove('hidden');
    modeCards.forEach(card => card.classList.toggle('active', card.dataset.mode === normalized));
    if (clientAlias) {
      const options = [...clientAlias.options];
      options.forEach(opt => {
        opt.hidden = normalized === 'taggd' ? opt.value !== 'Taggd' : opt.value === 'Taggd';
      });
      if (normalized === 'taggd') {
        clientAlias.value = 'Taggd';
      } else if (clientAlias.value === 'Taggd' || !clientAlias.value) {
        const fallback = options.find(opt => !opt.hidden && opt.value);
        if (fallback) clientAlias.value = fallback.value;
      }
      updateTaggdCandidateFields();
    }
    if (normalized === 'taggd') {
      prepareTaggdDefaults();
      ensureTaggdRows();
      taggdModal.showModal();
    }
  }
  modeCards.forEach(card => card.addEventListener('click', () => applyInvoiceMode(card.dataset.mode)));
  if (invoiceModeInput?.value) applyInvoiceMode(invoiceModeInput.value);
  document.getElementById('cancelCompany')?.addEventListener('click', () => {
    toggle.checked = false;
    Object.keys(map).forEach(name => document.querySelector(`[name="${name}"]`).value = '');
    modal.close();
  });
  document.getElementById('saveCompany')?.addEventListener('click', () => {
    Object.entries(map).forEach(([name, id]) => {
      document.querySelector(`[name="${name}"]`).value = document.getElementById(id).value.trim();
    });
    modal.close();
  });
  function taggdMoney(amount) {
    return new Intl.NumberFormat('en-IN', {style:'currency', currency:'INR', maximumFractionDigits:0}).format(Number(amount || 0));
  }
  function taggdKind(description) {
    const text = String(description || '').toLowerCase();
    if (text.includes('mark') && text.includes('up')) return 'markup';
    if (text.includes('laptop')) return 'laptop';
    if (text.includes('seat')) return 'seat';
    if (text.includes('recruiter') && text.includes('salar')) return 'recruiter_salary';
    if (text.includes('manager') && text.includes('salar')) return 'manager_salary';
    return 'other';
  }
  function updateTaggdCandidateFields() {
    const isTaggd = clientAlias?.value === 'Taggd';
    document.querySelectorAll('.candidate-field').forEach(field => {
      field.classList.toggle('hidden-for-taggd', isTaggd);
    });
    ['candidate_name', 'ctc', 'role', 'joining_date'].forEach(name => {
      const input = document.querySelector(`[name="${name}"]`);
      if (input) input.required = !isTaggd;
    });
  }
  function taggdRowTemplate(item = {}) {
    const row = document.createElement('div');
    row.className = 'taggd-row';
    row.innerHTML = `
      <label>Description<input class="taggd-desc" value="${String(item.description || '').replaceAll('"', '&quot;')}"></label>
      <label>Count<input class="taggd-count" type="number" step="0.01" min="0" value="${item.count ?? ''}"></label>
      <label>Rate<input class="taggd-rate" type="number" step="0.01" min="0" value="${item.rate || ''}"></label>
      <label>Amount<input class="taggd-amount" type="number" step="0.01" min="0" value="${item.amount || ''}"></label>`;
    row.querySelectorAll('input').forEach(input => input.addEventListener('input', recalcTaggdRows));
    return row;
  }
  function ensureTaggdRows() {
    if (!taggdRows.children.length) taggdDefaults.forEach(item => taggdRows.appendChild(taggdRowTemplate(item)));
    recalcTaggdRows();
  }
  function recalcTaggdRows() {
    let markupBase = 0;
    let markupRows = [];
    [...taggdRows.querySelectorAll('.taggd-row')].forEach(row => {
      const descInput = row.querySelector('.taggd-desc');
      const countInput = row.querySelector('.taggd-count');
      const rateInput = row.querySelector('.taggd-rate');
      const amountInput = row.querySelector('.taggd-amount');
      const kind = taggdKind(descInput.value);
      const count = Number(countInput.value || 0);
      let rate = Number(rateInput.value || 0);
      let amount = Number(amountInput.value || 0);
      if (kind === 'seat') {
        rate = 4500;
        amount = count * rate;
        rateInput.value = rate;
        amountInput.value = amount.toFixed(2);
        markupBase += amount;
      } else if (kind === 'laptop') {
        rate = 3333;
        amount = count * rate;
        rateInput.value = rate;
        amountInput.value = amount.toFixed(2);
      } else if (kind === 'recruiter_salary' || kind === 'manager_salary') {
        if (!amount && count && rate) {
          amount = count * rate;
          amountInput.value = amount.toFixed(2);
        }
        markupBase += amount;
      } else if (kind === 'markup') {
        countInput.value = 15;
        rateInput.value = '';
        markupRows.push(amountInput);
      } else if (!amount && count && rate) {
        amount = count * rate;
        amountInput.value = amount.toFixed(2);
      }
    });
    const markupAmount = markupBase * 0.15;
    markupRows.forEach(input => input.value = markupAmount.toFixed(2));
    updateTaggdSubtotal();
  }
  function readTaggdRows() {
    return [...taggdRows.querySelectorAll('.taggd-row')].map(row => {
      const count = Number(row.querySelector('.taggd-count').value || 0);
      const rate = Number(row.querySelector('.taggd-rate').value || 0);
      const amountInput = Number(row.querySelector('.taggd-amount').value || 0);
      return {
        description: row.querySelector('.taggd-desc').value.trim(),
        count,
        rate,
        amount: amountInput || (count * rate)
      };
    }).filter(item => item.description);
  }
  function updateTaggdSubtotal() {
    const subtotal = readTaggdRows().reduce((sum, item) => sum + Number(item.amount || 0), 0);
    document.getElementById('taggdSubtotal').textContent = `Subtotal: ${taggdMoney(subtotal)}`;
  }
  function prepareTaggdDefaults() {
    document.querySelector('[name="candidate_name"]').value ||= 'Taggd';
    document.querySelector('[name="role"]').value ||= 'Monthly Service Invoice';
    document.querySelector('[name="ctc"]').value ||= '0';
    document.querySelector('[name="joining_date"]').value ||= new Date().toISOString().slice(0, 10);
    const ref = document.querySelector('[name="reference"]');
    if (ref) ref.value = 'Amit Garg';
  }
  clientAlias?.addEventListener('change', () => {
    updateTaggdCandidateFields();
    if (clientAlias.value === 'Taggd') {
      if (invoiceModeInput) invoiceModeInput.value = 'taggd';
      if (invoiceFormWrap) invoiceFormWrap.classList.remove('hidden');
      modeCards.forEach(card => card.classList.toggle('active', card.dataset.mode === 'taggd'));
      prepareTaggdDefaults();
      ensureTaggdRows();
      taggdModal.showModal();
    } else if (invoiceModeInput && invoiceModeInput.value !== 'taggd' && invoiceModeInput.value !== 'non_taggd') {
      invoiceModeInput.value = 'non_taggd';
    }
  });
  updateTaggdCandidateFields();
  document.getElementById('addTaggdRow')?.addEventListener('click', () => {
    taggdRows.appendChild(taggdRowTemplate({description:'', count:1, rate:0, amount:0}));
    recalcTaggdRows();
  });
  document.getElementById('cancelTaggd')?.addEventListener('click', () => taggdModal.close());
  document.getElementById('saveTaggd')?.addEventListener('click', () => {
    recalcTaggdRows();
    taggdServiceItemsInput.value = JSON.stringify(readTaggdRows());
    taggdModal.close();
  });
  salaryReconForm?.addEventListener('submit', () => {
    if (reconcileTaggdServiceItems && taggdServiceItemsInput) {
      reconcileTaggdServiceItems.value = taggdServiceItemsInput.value || '';
    }
  });
  document.querySelector('form')?.addEventListener('submit', () => {
    if (invoiceModeInput && !invoiceModeInput.value) {
      invoiceModeInput.value = clientAlias?.value === 'Taggd' ? 'taggd' : 'non_taggd';
    }
    if (clientAlias?.value === 'Taggd') {
      prepareTaggdDefaults();
      ensureTaggdRows();
      recalcTaggdRows();
      taggdServiceItemsInput.value = JSON.stringify(readTaggdRows());
    }
  });
</script>
</body>
</html>
"""


app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
app.jinja_env.filters["money"] = indian_currency
app.jinja_env.globals["money"] = indian_currency
invoice_bp = Blueprint("invoice", __name__)


def invoice_endpoint(name: str) -> str:
    return f"{request.blueprint}.{name}" if request.blueprint else name


def invoice_route(rule: str, methods: list[str]):
    def decorator(func):
        app.add_url_rule(rule, func.__name__, func, methods=methods)
        invoice_bp.add_url_rule(rule, func.__name__, func, methods=methods)
        return func
    return decorator


def render_invoice_page(invoice=None, form=None, reconcile_result=None, selected_mode=None):
    mode = selected_mode or (form or {}).get("invoice_mode") or request.args.get("mode") or ""
    return render_template_string(
        HTML,
        clients=load_clients(),
        references=load_references(),
        company=COMPANY,
        invoice=invoice,
        form=form or {},
        next_invoice=next_invoice_number(),
        route=invoice_endpoint,
        money=indian_currency,
        taggd_defaults=TAGGD_DEFAULT_SERVICE_ITEMS,
        selected_mode=mode,
        reconcile_result=reconcile_result,
        reconcile_target=(reconcile_result or {}).get("target_total") if reconcile_result else None,
    )


def public_auth_enabled() -> bool:
    return bool(os.environ.get("INVOICE_APP_USERNAME") and os.environ.get("INVOICE_APP_PASSWORD"))


@app.before_request
def require_public_auth():
    if not public_auth_enabled():
        return None
    auth = request.authorization
    expected_user = os.environ["INVOICE_APP_USERNAME"]
    expected_password = os.environ["INVOICE_APP_PASSWORD"]
    if (
        auth
        and hmac.compare_digest(auth.username or "", expected_user)
        and hmac.compare_digest(auth.password or "", expected_password)
    ):
        return None
    return Response(
        "Authentication required",
        401,
        {"WWW-Authenticate": 'Basic realm="HR Guru Invoice App"'},
    )


@invoice_route("/", ["GET"])
def index():
    return render_invoice_page()


@invoice_route("/generate", ["POST"])
def generate():
    form = request.form.copy()
    invoice_mode = (form.get("invoice_mode") or "").strip().lower()
    if invoice_mode == "taggd":
        form["client_alias"] = "Taggd"
    elif invoice_mode == "non_taggd" and form.get("client_alias") == "Taggd":
        non_taggd_alias = next((alias for alias in load_clients().keys() if alias != "Taggd"), "")
        if non_taggd_alias:
            form["client_alias"] = non_taggd_alias
    new_client_alias = save_client(form)
    if new_client_alias and form.get("new_client_name", "").strip():
        form["client_alias"] = new_client_alias
    new_reference = save_reference(form.get("new_reference", ""))
    if new_reference:
        form["reference"] = new_reference
    invoice = build_invoice(form)
    row = asdict(invoice)
    row["client"] = invoice.client_alias
    save_generated(row)
    pdf_for(invoice)
    return render_invoice_page(invoice=invoice, form=form)


@invoice_route("/reconcile-salary-slips", ["POST"])
def reconcile_salary_slips_route():
    form = request.form.copy()
    form["invoice_mode"] = "taggd"
    target_total = taggd_recruiter_salary_total(form)
    files = request.files.getlist("salary_slips")
    result = reconcile_salary_slips(files, target_total)
    if not result["target_total"]:
        result["note"] = "No Taggd recruiter salary total could be derived from the current service details."
    return render_invoice_page(form=form, reconcile_result=result, selected_mode="taggd")


@invoice_route("/logo.png", ["GET"])
def logo():
    if LOGO_PATH.exists():
        return send_file(LOGO_PATH)
    return Response(status=404)


@invoice_route("/invoice/<invoice_number>.pdf", ["GET"])
def download_pdf(invoice_number: str):
    matches = sorted(OUTPUT_DIR.glob(f"{invoice_number}_*.pdf"))
    if not matches and LEGACY_OUTPUT_DIR.exists():
        matches = sorted(LEGACY_OUTPUT_DIR.glob(f"{invoice_number}_*.pdf"))
    if not matches:
        return redirect(url_for(invoice_endpoint("index")))
    return send_file(matches[-1], as_attachment=True, download_name=matches[-1].name)


@invoice_route("/mis.json", ["GET"])
def mis():
    return Response(json.dumps(load_mis(), indent=2), mimetype="application/json")


@invoice_route("/mis.xlsx", ["GET"])
def export_mis():
    path = export_mis_excel()
    return send_file(path, as_attachment=True, download_name="Invoice_MIS_Export.xlsx")


if __name__ == "__main__":
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    app.run(
        host=os.environ.get("HOST", "127.0.0.1"),
        port=int(os.environ.get("PORT", "5055")),
        debug=os.environ.get("FLASK_DEBUG") == "1",
    )
